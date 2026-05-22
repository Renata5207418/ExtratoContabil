import io
import os
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from dateutil.relativedelta import relativedelta
from bson import ObjectId
from services.mongo import clientes_col, arquivos_col, solicitacoes_col, usuarios_col
from services.db_dominio import get_todas_empresas_dominio


extratos_bp = Blueprint('extratos', __name__, url_prefix='/api/extratos')


def formatar_data_hora(data):
    if isinstance(data, datetime):
        return data.strftime("%d/%m/%Y %H:%M")
    return None


def formatar_historico_edicoes(historico):
    historico_formatado = []

    for item in historico[-50:]:
        historico_formatado.append({
            "usuario_id": item.get("usuario_id"),
            "usuario_nome": item.get("usuario_nome", "Sistema"),
            "data": formatar_data_hora(item.get("data")),
            "alteracoes": item.get("alteracoes", {})
        })

    return historico_formatado


# ==========================================
# 1. TELA DE ACOMPANHAMENTO 
# ==========================================
@extratos_bp.route('/acompanhamento', methods=['GET'])
@jwt_required()
def listar_acompanhamento():
    mes_padrao = (datetime.now() - relativedelta(months=1)).strftime("%m.%Y")  
    mes_filtro = request.args.get('mes', mes_padrao)
    
    empresas = list(clientes_col.find({"ativo": True}))
    resultado = []
    
    for emp in empresas:
        cliente_id = emp["_id"]
        
        solicitacao = solicitacoes_col.find_one({
            "cliente_id": cliente_id,
            "mes_referencia": mes_filtro
        })
        
        qtd_recebidos = 0
        qtd_processados = 0
        data_recente = "-"
        validado = False
        
        if solicitacao:
            arquivos = list(arquivos_col.find({"solicitacao_id": solicitacao["_id"]}))
            qtd_recebidos = len(arquivos)
            qtd_processados = sum(1 for arq in arquivos if arq.get("conteudo") and arq.get("conteudo") != "Resumo indisponível")
            validado = solicitacao.get("validado", False)
            
            if qtd_recebidos > 0:
                datas = [arq.get("data_leitura") for arq in arquivos if arq.get("data_leitura")]
                if datas:
                    data_recente = max(datas).strftime("%d/%m/%Y")
        
        resultado.append({
            "codigo": emp.get("codigo_dominio"),
            "id_solicitacao": str(solicitacao["_id"]) if solicitacao else None, 
            "cnpj": emp.get("cnpj", ""),
            "empresa": emp.get("nome_empresa"),
            "recebidos": qtd_recebidos,
            "processados": qtd_processados,
            "data": data_recente,
            "validado": validado,
            "validado_por": solicitacao.get("validado_por") if solicitacao else None,
            "data_validacao": solicitacao.get("data_validacao").strftime("%d/%m/%Y %H:%M") if solicitacao and solicitacao.get("data_validacao") else None
            
        })
        
    resultado.sort(key=lambda x: x["codigo"])

    return jsonify({"mes_referencia": mes_filtro, "dados": resultado}), 200


# ==========================================
# 2. TELA DETALHE CLIENTE 
# ==========================================
@extratos_bp.route('/cliente/<int:codigo_dominio>', methods=['GET'])
@jwt_required()
def detalhe_cliente(codigo_dominio):
    mes_filtro = request.args.get('mes', datetime.now().strftime("%m.%Y"))
    
    cliente = clientes_col.find_one({"codigo_dominio": codigo_dominio})
    if not cliente:
        return jsonify({"erro": "Cliente não encontrado."}), 404
        
    solicitacao = solicitacoes_col.find_one({
        "cliente_id": cliente["_id"],
        "mes_referencia": mes_filtro
    })
    
    lista_arquivos = []
    info_solicitacao = None
    
    if solicitacao:
        info_solicitacao = {
            "id": str(solicitacao["_id"]),
            "validado": solicitacao.get("validado", False),
            "validado_por": solicitacao.get("validado_por"),
            "data_validacao": solicitacao.get("data_validacao").strftime("%d/%m/%Y %H:%M") if solicitacao.get("data_validacao") else None
        }

        arquivos = arquivos_col.find({"solicitacao_id": solicitacao["_id"]})
        
        for arq in arquivos:
            banco = arq.get("banco") or "Inexistente"
            periodo = arq.get("periodo") or "Não identificado"
            banco_periodo = f"{banco} / {periodo}" if arq.get("banco") else periodo
            
            dt_leitura = arq.get("data_leitura")
            data_formatada = dt_leitura.strftime("%d/%m/%Y %H:%M") if dt_leitura else "-"
            
            lista_arquivos.append({
                "id": str(arq["_id"]),
                "arquivo": arq.get("nome_arquivo"),
                "tipo": arq.get("tipo_documento", "Desconhecido").title(),
                "banco_periodo": banco_periodo,
                "data": data_formatada,
                "resumo_detalhado": arq.get("conteudo", "Resumo indisponível"),
                "numero_solicitacao": arq.get("numero_solicitacao"),
                "observacao": arq.get("observacao", ""),

                "status": arq.get("status", ""),
                "status_leitura": arq.get("status_leitura", ""),
                "mensagem_leitura": arq.get("mensagem_leitura", ""),

                "ultima_edicao_por": arq.get("ultima_edicao_por"),
                "ultima_edicao_em": formatar_data_hora(arq.get("ultima_edicao_em")),
                "historico_edicoes": formatar_historico_edicoes(
                    arq.get("historico_edicoes", [])
                )
            })
            
    return jsonify({
        "cliente": {
            "codigo": cliente.get("codigo_dominio"),
            "cnpj": cliente.get("cnpj"),
            "nome": cliente.get("nome_empresa")
        },
        "solicitacao": info_solicitacao,
        "arquivos": lista_arquivos
    }), 200


# ==========================================
# 3. EXPORTAR RELATÓRIO EM XLSX FORMATADO
# ==========================================
@extratos_bp.route('/solicitacao/<solicitacao_id>/exportar', methods=['GET'])
@jwt_required()
def exportar_relatorio_solicitacao(solicitacao_id):
    try:
        solicitacao = solicitacoes_col.find_one({"_id": ObjectId(solicitacao_id)})
        if not solicitacao:
            return jsonify({"erro": "Solicitação não encontrada"}), 404

        cliente = clientes_col.find_one({"_id": solicitacao["cliente_id"]})
        nome_empresa = cliente.get("nome_empresa", "Desconhecido") if cliente else "Desconhecido"
        codigo_dominio = cliente.get("codigo_dominio", "N/A") if cliente else "N/A"
        cnpj_empresa = cliente.get("cnpj", "N/A") if cliente else "N/A"
        mes_referencia = solicitacao.get("mes_referencia", "")

        arquivos = list(arquivos_col.find({"solicitacao_id": ObjectId(solicitacao_id)}))

        def formatar_cnpj(cnpj):
            somente_numeros = ''.join(filter(str.isdigit, str(cnpj or "")))

            if len(somente_numeros) == 14:
                return (
                    f"{somente_numeros[:2]}."
                    f"{somente_numeros[2:5]}."
                    f"{somente_numeros[5:8]}/"
                    f"{somente_numeros[8:12]}-"
                    f"{somente_numeros[12:]}"
                )

            return str(cnpj or "N/A")

        # Cria planilha
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório"

        # Cores
        cinza_claro = "E7E6E6"
        cinza_titulo = "D9D9D9"
        cinza_borda = "BFBFBF"
        branco = "FFFFFF"

        # Estilos
        fonte_titulo = Font(bold=True, size=14, color="000000")
        fonte_subtitulo = Font(bold=True, size=11, color="000000")
        fonte_header = Font(bold=True, color="000000")
        fonte_normal = Font(size=10, color="000000")

        fill_titulo = PatternFill("solid", fgColor=cinza_titulo)
        fill_header = PatternFill("solid", fgColor=cinza_claro)

        borda_fina = Border(
            left=Side(style="thin", color=cinza_borda),
            right=Side(style="thin", color=cinza_borda),
            top=Side(style="thin", color=cinza_borda),
            bottom=Side(style="thin", color=cinza_borda),
        )

        alinhamento_centro = Alignment(horizontal="center", vertical="center")
        alinhamento_esquerda = Alignment(horizontal="left", vertical="center")
        alinhamento_quebra = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Título principal
        ws.merge_cells("A1:D1")
        ws["A1"] = "RELATÓRIO DE EXTRATOS"
        ws["A1"].font = fonte_titulo
        ws["A1"].fill = fill_titulo
        ws["A1"].alignment = alinhamento_centro
        ws["A1"].border = borda_fina
        ws.row_dimensions[1].height = 26

        # Informações da empresa
        dados_empresa = [
            ("EMPRESA", nome_empresa),
            ("CÓDIGO DOMÍNIO", str(codigo_dominio)),
            ("CNPJ", formatar_cnpj(cnpj_empresa)),
            ("COMPETÊNCIA", str(mes_referencia)),
        ]

        linha = 3
        for label, valor in dados_empresa:
            ws[f"A{linha}"] = label
            ws[f"B{linha}"] = valor

            ws[f"A{linha}"].font = fonte_subtitulo
            ws[f"A{linha}"].fill = fill_header
            ws[f"A{linha}"].alignment = alinhamento_esquerda
            ws[f"A{linha}"].border = borda_fina

            ws[f"B{linha}"].font = fonte_normal
            ws[f"B{linha}"].alignment = alinhamento_esquerda
            ws[f"B{linha}"].border = borda_fina
            ws[f"B{linha}"].number_format = "@"

            # Mescla B até D para ficar mais bonito
            ws.merge_cells(start_row=linha, start_column=2, end_row=linha, end_column=4)

            linha += 1

        # Cabeçalho da tabela
        linha_header = 8
        headers = ["NOME DO EXTRATO", "BANCO", "PERÍODO", "OBSERVAÇÃO"]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=linha_header, column=col_idx)
            cell.value = header
            cell.font = fonte_header
            cell.fill = fill_header
            cell.border = borda_fina
            cell.alignment = alinhamento_centro

        # Dados dos arquivos
        linha_atual = linha_header + 1

        for arq in arquivos:
            dados_linha = [
                arq.get("nome_arquivo", ""),
                arq.get("banco", "Desconhecido"),
                arq.get("periodo", "N/A"),
                arq.get("observacao", ""),
            ]

            for col_idx, valor in enumerate(dados_linha, start=1):
                cell = ws.cell(row=linha_atual, column=col_idx)
                cell.value = str(valor or "")
                cell.font = fonte_normal
                cell.border = borda_fina
                cell.alignment = alinhamento_quebra
                cell.number_format = "@"

            linha_atual += 1

        # Caso não tenha arquivos
        if not arquivos:
            ws.cell(row=linha_atual, column=1).value = "Nenhum arquivo encontrado para esta solicitação."
            ws.cell(row=linha_atual, column=1).font = fonte_normal
            ws.cell(row=linha_atual, column=1).alignment = alinhamento_esquerda
            ws.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=4)

        # Larguras das colunas
        larguras = {
            "A": 55,
            "B": 24,
            "C": 16,
            "D": 45,
        }

        for col, largura in larguras.items():
            ws.column_dimensions[col].width = largura

        # Altura das linhas da tabela
        for row in range(linha_header + 1, linha_atual + 1):
            ws.row_dimensions[row].height = 24

        # Congelar cabeçalho da tabela
        ws.freeze_panes = "A9"

        # Filtro na tabela
        ultima_linha = max(linha_atual - 1, linha_header)
        ws.auto_filter.ref = f"A{linha_header}:D{ultima_linha}"

        # Ajuste visual geral
        ws.sheet_view.showGridLines = False

        # Gera arquivo em memória
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Relatorio_Extratos_{codigo_dominio}_{mes_referencia}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return jsonify({"erro": str(e)}), 500
    


# ==========================================
# 4. SALVAR DETALHES DO ARQUIVO COM HISTÓRICO
# ==========================================
@extratos_bp.route('/arquivo/<arquivo_id>/detalhes', methods=['PUT'])
@jwt_required()
def atualizar_detalhes_arquivo(arquivo_id):
    try:
        if not ObjectId.is_valid(arquivo_id):
            return jsonify({"erro": "ID de arquivo inválido."}), 400

        arquivo = arquivos_col.find_one({"_id": ObjectId(arquivo_id)})

        if not arquivo:
            return jsonify({"erro": "Arquivo não encontrado."}), 404

        dados = request.json or {}

        user_id = get_jwt_identity()
        user = usuarios_col.find_one({"_id": ObjectId(user_id)}) if ObjectId.is_valid(str(user_id)) else None
        username = user.get("username", "Sistema") if user else "Sistema"

        campos_permitidos = {
            "banco": "Banco",
            "periodo": "Período",
            "observacao": "Observação"
        }

        alteracoes = {}
        dados_update = {}

        for campo, label in campos_permitidos.items():
            if campo not in dados:
                continue

            novo_valor = dados.get(campo)

            if novo_valor is None:
                novo_valor = ""

            if campo in ["banco", "periodo"]:
                novo_valor = str(novo_valor).strip()
            else:
                novo_valor = str(novo_valor)

            valor_antigo = arquivo.get(campo) or ""

            if str(valor_antigo) != str(novo_valor):
                alteracoes[campo] = {
                    "label": label,
                    "antes": str(valor_antigo),
                    "depois": str(novo_valor)
                }

                dados_update[campo] = novo_valor

        if not alteracoes:
            return jsonify({
                "mensagem": "Nenhuma alteração identificada.",
                "arquivo": {
                    "id": str(arquivo["_id"]),
                    "banco": arquivo.get("banco", "Desconhecido"),
                    "periodo": arquivo.get("periodo", "N/A"),
                    "observacao": arquivo.get("observacao", ""),
                    "banco_periodo": f"{arquivo.get('banco', 'Desconhecido')} / {arquivo.get('periodo', 'N/A')}",
                    "ultima_edicao_por": arquivo.get("ultima_edicao_por"),
                    "ultima_edicao_em": formatar_data_hora(arquivo.get("ultima_edicao_em")),
                    "historico_edicoes": formatar_historico_edicoes(
                        arquivo.get("historico_edicoes", [])
                    )
                }
            }), 200

        agora = datetime.now()

        registro_historico = {
            "usuario_id": str(user_id),
            "usuario_nome": username,
            "data": agora,
            "alteracoes": alteracoes
        }

        dados_update.update({
            "ultima_edicao_por": username,
            "ultima_edicao_por_id": str(user_id),
            "ultima_edicao_em": agora
        })

        arquivos_col.update_one(
            {"_id": ObjectId(arquivo_id)},
            {
                "$set": dados_update,
                "$push": {
                    "historico_edicoes": {
                        "$each": [registro_historico],
                        "$slice": -100
                    }
                }
            }
        )

        arquivo_atualizado = arquivos_col.find_one({"_id": ObjectId(arquivo_id)})

        banco = arquivo_atualizado.get("banco", "Desconhecido")
        periodo = arquivo_atualizado.get("periodo", "N/A")

        return jsonify({
            "mensagem": "Informações atualizadas com histórico.",
            "arquivo": {
                "id": str(arquivo_atualizado["_id"]),
                "banco": banco,
                "periodo": periodo,
                "observacao": arquivo_atualizado.get("observacao", ""),
                "banco_periodo": f"{banco} / {periodo}",
                "ultima_edicao_por": arquivo_atualizado.get("ultima_edicao_por"),
                "ultima_edicao_em": formatar_data_hora(
                    arquivo_atualizado.get("ultima_edicao_em")
                ),
                "historico_edicoes": formatar_historico_edicoes(
                    arquivo_atualizado.get("historico_edicoes", [])
                )
            }
        }), 200

    except Exception as e:
        return jsonify({"erro": str(e)}), 500


# ==========================================
# 4.1. SALVAR EDIÇÕES MANUAIS NO ARQUIVO
# ==========================================
@extratos_bp.route('/arquivo/<arquivo_id>/<campo>', methods=['PUT'])
@jwt_required()
def atualizar_arquivo(arquivo_id, campo):
    campos_permitidos = ['banco', 'periodo', 'observacao']
    
    if campo not in campos_permitidos:
        return jsonify({"erro": "Campo inválido para atualização"}), 400
        
    dados = request.json
    novo_valor = dados.get(campo)
    
    try:
        resultado = arquivos_col.update_one(
            {"_id": ObjectId(arquivo_id)},
            {"$set": {campo: novo_valor}}
        )
        
        if resultado.matched_count == 0:
            return jsonify({"erro": "Arquivo não encontrado."}), 404
            
        return jsonify({"mensagem": f"{campo.capitalize()} atualizado com sucesso!"}), 200
    except Exception as e:
        return jsonify({"erro": str(e)}), 500


# ==========================================
# 4.2. VISUALIZAR PDF ORIGINAL
# ==========================================
@extratos_bp.route('/arquivo/<arquivo_id>/visualizar', methods=['GET'])
@jwt_required()
def visualizar_arquivo_original(arquivo_id):
    try:
        if not ObjectId.is_valid(arquivo_id):
            return jsonify({"erro": "ID de arquivo inválido."}), 400

        arquivo = arquivos_col.find_one({"_id": ObjectId(arquivo_id)})

        if not arquivo:
            return jsonify({"erro": "Arquivo não encontrado."}), 404

        caminho = arquivo.get("caminho_completo")
        nome_arquivo = arquivo.get("nome_arquivo", "arquivo.pdf")

        if not caminho:
            return jsonify({"erro": "Caminho do arquivo não encontrado no banco."}), 404

        caminho_absoluto = os.path.abspath(caminho)

        if not os.path.exists(caminho_absoluto):
            return jsonify({
                "erro": "Arquivo não encontrado na rede.",
                "arquivo": nome_arquivo
            }), 404

        if not caminho_absoluto.lower().endswith(".pdf"):
            return jsonify({"erro": "O arquivo não é um PDF válido."}), 400

        return send_file(
            caminho_absoluto,
            mimetype="application/pdf",
            as_attachment=False,
            download_name=nome_arquivo,
            conditional=True
        )

    except Exception as e:
        return jsonify({"erro": str(e)}), 500
    

# ==========================================
# 5. VALIDAÇÃO DE LOTE
# ==========================================
@extratos_bp.route('/validar/<solicitacao_id>', methods=['POST'])
@jwt_required()
def validar_lote(solicitacao_id):
    user_id = get_jwt_identity() 
    user = usuarios_col.find_one({"_id": ObjectId(user_id)})
    
    if not user:
        return jsonify({"erro": "Usuário não encontrado."}), 404
        
    username = user.get("username", "Usuário")
    
    resultado = solicitacoes_col.update_one(
        {"_id": ObjectId(solicitacao_id)},
        {"$set": {
            "validado": True,
            "validado_por": username,
            "data_validacao": datetime.now()
        }}
    )
    
    if resultado.modified_count == 0:
        return jsonify({"erro": "Solicitação não encontrada."}), 404
        
    return jsonify({"mensagem": "Lote validado com sucesso!", "validado_por": username}), 200

# ==========================================
# 6. DADOS DO DASHBOARD
# ==========================================
@extratos_bp.route('/dashboard', methods=['GET'])
@jwt_required()
def dados_dashboard():
    mes_padrao = (datetime.now() - relativedelta(months=1)).strftime("%m.%Y")
    mes_filtro = request.args.get('mes', mes_padrao)

    empresas_dominio = get_todas_empresas_dominio()
    total_carteira = len(empresas_dominio)

    sols_mes = list(solicitacoes_col.find({"mes_referencia": mes_filtro}))
    sols_map = {str(s["cliente_id"]): s for s in sols_mes}

    clientes_mongo = list(clientes_col.find({}))
    cliente_id_por_codigo = {c["codigo_dominio"]: str(c["_id"]) for c in clientes_mongo}

    empresas_com_envio = []
    empresas_sem_envio = []
    total_pdfs = 0

    for emp in empresas_dominio:
        codigo = emp["codigo_dominio"]
        cliente_mongo_id = cliente_id_por_codigo.get(codigo)

        qtd_recebidos = 0
        if cliente_mongo_id and cliente_mongo_id in sols_map:
            solicitacao = sols_map[cliente_mongo_id]
            qtd_recebidos = arquivos_col.count_documents({"solicitacao_id": solicitacao["_id"]})

        dados_empresa = {
            "Código Domínio": codigo,
            "CNPJ": emp["cnpj"],
            "Razão Social": emp["nome_empresa"],
            "Arquivos Processados": qtd_recebidos
        }

        if qtd_recebidos > 0:
            empresas_com_envio.append(dados_empresa)
            total_pdfs += qtd_recebidos
        else:
            empresas_sem_envio.append(dados_empresa)

    grafico = []
    for i in range(4, -1, -1):
        mes_grafico = (datetime.now() - relativedelta(months=i+1))
        mes_str = mes_grafico.strftime("%m.%Y")
        mes_nome = mes_grafico.strftime("%b").capitalize()

        sols_periodo = list(solicitacoes_col.find({"mes_referencia": mes_str}))
        sols_ids = [s["_id"] for s in sols_periodo]
        qtd_mes = arquivos_col.count_documents({"solicitacao_id": {"$in": sols_ids}})
        grafico.append({"name": mes_nome, "envios": qtd_mes})

    engajamento_pct = round((len(empresas_com_envio) / total_carteira * 100), 1) if total_carteira > 0 else 0
    pendente_pct = round((len(empresas_sem_envio) / total_carteira * 100), 1) if total_carteira > 0 else 0

    return jsonify({
        "estatisticas": {
            "carteira": total_carteira,
            "processamento": total_pdfs,
            "engajamento": len(empresas_com_envio),
            "engajamento_pct": engajamento_pct,
            "pendente": len(empresas_sem_envio),
            "pendente_pct": pendente_pct
        },
        "listas_excel": {
            "com_envio": empresas_com_envio,
            "sem_envio": empresas_sem_envio
        },
        "grafico": grafico
    }), 200


@extratos_bp.route('/toggle-validar/<solicitacao_id>', methods=['POST'])
@jwt_required()
def toggle_validar(solicitacao_id):
    user_id = get_jwt_identity()
    user = usuarios_col.find_one({"_id": ObjectId(user_id)})
    username = user.get("username", "Sistema") if user else "Sistema"
    
    solicitacao = solicitacoes_col.find_one({"_id": ObjectId(solicitacao_id)})
    if not solicitacao:
        return jsonify({"erro": "Solicitação não encontrada."}), 404
        
    novo_status = not solicitacao.get("validado", False)
    
    update_data = {
        "validado": novo_status,
        "validado_por": username if novo_status else None,
        "data_validacao": datetime.now() if novo_status else None
    }
    
    solicitacoes_col.update_one({"_id": ObjectId(solicitacao_id)}, {"$set": update_data})
    
    return jsonify({"mensagem": "Status alterado!", "validado": novo_status}), 200

